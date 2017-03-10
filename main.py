from openpyxl import Workbook
from openpyxl import load_workbook

if __name__ == '__main__':
    excel_data = Workbook()
    excel_data = load_workbook("C:/Users/Owner/Documents/GitHub/TATADataChallenge2017/NBTC_Tata_Challenge.01.xlsx")
    #excel_data.create_sheet("video_game_data", 0)
    print(excel_data.get_sheet_names)
